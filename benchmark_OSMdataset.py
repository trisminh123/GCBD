"""
benchmark_realdataset.py
So sánh GCBD original vs GCBD Gaussian vs CLIQUE
Chạy trên các dataset dạng file .txt (tab-separated) có cột PoiID/IdPoints, NEAR_X, NEAR_Y

Chạy:
    python benchmark_realdataset.py

Kết quả:
    realworld_dataset_results/results.xlsx  ← file Excel theo mẫu
"""

import os
import time
import warnings
import itertools
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.metrics import (
    silhouette_score,
    davies_bouldin_score,
    calinski_harabasz_score,
)
from sklearn.preprocessing import MinMaxScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from Algorithm.fastgrid import fastgrid
from Algorithm.fastgrid_gaussian import fastgrid_gaussian

try:
    from pyclustering.cluster.clique import clique
    CLIQUE_AVAILABLE = True
except ImportError:
    CLIQUE_AVAILABLE = False
    print("CẢNH BÁO: pyclustering chưa được cài. Chạy: pip install pyclustering")


# ─────────────────────────────────────────────
# CẤU HÌNH
# ─────────────────────────────────────────────
RESULT_DIR = "results/OSM"

# Thư mục chứa các file .txt dataset
DATASET_DIR = "Datasets/OSM_dataset"   

# Grid search — cùng không gian tham số cho GCBD gốc và Gaussian
GRID_PARAMS = {
    "no_grid"   : list(range(6, 52)),   # l = 5 → 50
    "percentile": [0.1],
    "max_iters" : list(range(2, 13)),   # T = 2 → 12
}

# Gaussian thêm sigma (đồng bộ với file synthetic)
SIGMA_VALUES = [0.25, 0.5, 0.75, 1.0]

# CLIQUE grid search (đồng bộ với file synthetic)
CLIQUE_PARAMS = {
    "intervals": list(range(5, 51)),    # l = [5, 50]
    "threshold": list(range(0, 6)),     # c = [0, 1, 2, 3, 4, 5]
}
# ─────────────────────────────────────────────


def load_txt_dataset(filepath: str):
    """
    Đọc file .txt tab-separated.
    Cột ID có thể tên là 'PoiID' hoặc 'IdPoints'.
    Trả về mảng numpy shape (n, 2) gồm [NEAR_X, NEAR_Y].
    """
    df = pd.read_csv(filepath, sep="\t")

    # Chuẩn hóa tên cột
    df.columns = [c.strip() for c in df.columns]

    # Xử lý cột ID (PoiID hoặc IdPoints)
    id_col = None
    for candidate in ["PoiID", "IdPoints"]:
        if candidate in df.columns:
            id_col = candidate
            break
    if id_col is None:
        raise ValueError(f"Không tìm thấy cột PoiID hoặc IdPoints trong {filepath}")

    X = df[["NEAR_X", "NEAR_Y"]].values.astype(float)
    dataset_name = Path(filepath).stem
    return dataset_name, X


def compute_bss_wss(X: np.ndarray, labels: np.ndarray):
    """Tính BSS và WSS từ dữ liệu và nhãn cụm."""
    unique_labels = np.unique(labels[labels > 0])
    grand_mean = X.mean(axis=0)

    bss = 0.0
    wss = 0.0
    for lbl in unique_labels:
        mask = labels == lbl
        cluster_points = X[mask]
        cluster_mean = cluster_points.mean(axis=0)
        n_k = mask.sum()
        bss += n_k * np.sum((cluster_mean - grand_mean) ** 2)
        wss += np.sum((cluster_points - cluster_mean) ** 2)

    return bss, wss


def dunn_index(X: np.ndarray, labels: np.ndarray):
    """Tính Dunn Index."""
    unique_labels = np.unique(labels[labels > 0])
    if len(unique_labels) < 2:
        return 0.0

    clusters = [X[labels == lbl] for lbl in unique_labels]

    # Đường kính trong cụm (max pairwise distance inside cluster)
    max_intra = []
    for c in clusters:
        if len(c) < 2:
            max_intra.append(0.0)
        else:
            from sklearn.metrics import pairwise_distances
            d = pairwise_distances(c)
            max_intra.append(d.max())

    # Khoảng cách nhỏ nhất giữa hai cụm khác nhau
    min_inter = np.inf
    for i in range(len(clusters)):
        for j in range(i + 1, len(clusters)):
            from sklearn.metrics import pairwise_distances
            d = pairwise_distances(clusters[i], clusters[j])
            min_inter = min(min_inter, d.min())

    max_diam = max(max_intra) if max_intra else 1.0
    if max_diam == 0:
        return 0.0
    return min_inter / max_diam


def wb_index(wss: float, bss: float, n_clusters: int, n_samples: int):
    """Chỉ số WB = (n_clusters * WSS) / BSS."""
    if bss == 0:
        return np.nan
    return (n_clusters * wss) / bss


def compute_all_metrics(X, pred):
    """Tính toàn bộ chỉ số từ dữ liệu X và nhãn pred."""
    active = pred > 0
    if active.sum() < 2:
        return None

    X_active = X[active]
    labels_active = pred[active]
    unique = np.unique(labels_active)
    n_clusters = len(unique)

    if n_clusters < 2:
        return None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        sil = silhouette_score(X_active, labels_active)
        db  = davies_bouldin_score(X_active, labels_active)
        ch  = calinski_harabasz_score(X_active, labels_active)

    bss, wss = compute_bss_wss(X_active, labels_active)
    wb  = wb_index(wss, bss, n_clusters, len(X_active))
    dun = dunn_index(X_active, labels_active)

    return {
        "n_clusters" : n_clusters,
        "n_assigned" : int(active.sum()),
        "Silhouette" : round(sil, 6),
        "BSS"        : round(bss, 6),
        "WSS"        : round(wss, 6),
        "WB"         : round(wb,  6),
        "DB"         : round(db,  6),
        "Dunn"       : round(dun, 6),
        "CH"         : round(ch,  6),
    }


def clique_predict(data: np.ndarray, intervals: int, threshold: int):
    """
    Chạy CLIQUE và trả về mảng nhãn cluster (1-indexed, 0 = noise).
    Quy ước nhãn > 0 là điểm được phân cụm, giống với GCBD.
    """
    if not CLIQUE_AVAILABLE:
        return None

    # CLIQUE yêu cầu dữ liệu trong [0, 1]
    scaler = MinMaxScaler()
    data_scaled = scaler.fit_transform(data).tolist()

    clique_instance = clique(data_scaled, intervals, threshold)
    clique_instance.process()
    clusters = clique_instance.get_clusters()

    # Gán nhãn 1-indexed; điểm không thuộc cụm nào → 0 (noise)
    labels = np.zeros(len(data), dtype=int)
    for cluster_id, indices in enumerate(clusters, start=1):
        for idx in indices:
            labels[idx] = cluster_id
    return labels


def run_best_for_dataset(dataset_name, X):
    """
    Chạy grid search cho 3 thuật toán: GCBD gốc, GCBD Gaussian, CLIQUE.
    Trả về dict chứa best row cho mỗi thuật toán (tối ưu Silhouette).
    """
    print(f"\n{'='*60}")
    print(f"▶ Dataset: {dataset_name}  (n={len(X)})")
    print(f"{'='*60}")

    keys   = list(GRID_PARAMS.keys())
    values = list(GRID_PARAMS.values())
    best   = {"original": None, "gaussian": None, "clique": None}

    def try_update_best(key, metrics, elapsed, extra):
        if metrics is None:
            return
        if best[key] is None or metrics["Silhouette"] > best[key]["Silhouette"]:
            best[key] = {**metrics, "time": elapsed, **extra}

    n_combos = len(GRID_PARAMS["no_grid"]) * len(GRID_PARAMS["percentile"]) * len(GRID_PARAMS["max_iters"])
    n_combos_clique = len(CLIQUE_PARAMS["intervals"]) * len(CLIQUE_PARAMS["threshold"])

    # ── GCBD original ──────────────────────────────────────────────────────────
    print(f"  ▶ Chạy GCBD gốc ({n_combos} combos)...")
    for combo in itertools.product(*values):
        params = dict(zip(keys, combo))
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                t0   = time.perf_counter()
                pred = fastgrid(X,
                                no_grid    = params["no_grid"],
                                percentile = params["percentile"],
                                max_iters  = params["max_iters"])
                elapsed = time.perf_counter() - t0
        except Exception:
            continue
        metrics = compute_all_metrics(X, pred)
        try_update_best("original", metrics, elapsed,
                        {"no_grid"  : params["no_grid"],
                         "max_iters": params["max_iters"]})

    # ── GCBD Gaussian — từng sigma ─────────────────────────────────────────────
    for sigma in SIGMA_VALUES:
        print(f"  ▶ Chạy GCBD Gaussian σ={sigma} ({n_combos} combos)...")
        for combo in itertools.product(*values):
            params = dict(zip(keys, combo))
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    t0   = time.perf_counter()
                    pred = fastgrid_gaussian(X,
                                            no_grid    = params["no_grid"],
                                            percentile = params["percentile"],
                                            max_iters  = params["max_iters"],
                                            sigma      = sigma)
                    elapsed = time.perf_counter() - t0
            except Exception:
                continue
            metrics = compute_all_metrics(X, pred)
            # Cập nhật nếu tốt hơn gaussian hiện tại
            if metrics is None:
                continue
            if best["gaussian"] is None or metrics["Silhouette"] > best["gaussian"]["Silhouette"]:
                best["gaussian"] = {**metrics, "time": elapsed,
                                    "no_grid"  : params["no_grid"],
                                    "max_iters": params["max_iters"],
                                    "sigma"    : sigma}

    # ── CLIQUE ─────────────────────────────────────────────────────────────────
    if CLIQUE_AVAILABLE:
        print(f"  ▶ Chạy CLIQUE ({n_combos_clique} combos)...")
        for intervals in CLIQUE_PARAMS["intervals"]:
            for threshold in CLIQUE_PARAMS["threshold"]:
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        t0   = time.perf_counter()
                        pred = clique_predict(X, intervals, threshold)
                        elapsed = time.perf_counter() - t0
                    if pred is None:
                        continue
                except Exception:
                    continue
                metrics = compute_all_metrics(X, pred)
                try_update_best("clique", metrics, elapsed,
                                {"intervals": intervals,
                                 "threshold": threshold})
    else:
        print("  ▶ CLIQUE: bỏ qua (pyclustering chưa cài)")

    # In kết quả tốt nhất
    for key, label in [("original", "GCBD gốc"),
                        ("gaussian", "GCBD cải tiến"),
                        ("clique",   "CLIQUE")]:
        r = best[key]
        if r:
            sigma_str = f", σ={r['sigma']}" if key == "gaussian" else ""
            print(f"  📊 {label}: Sil={r['Silhouette']:.4f}, "
                  f"k={r['n_clusters']}{sigma_str}, t={r.get('time', 0):.2f}s")
        else:
            print(f"  📊 {label}: không có kết quả")

    return best


def write_xlsx(all_results: list, out_path: str):
    """
    Ghi kết quả vào file xlsx.
    all_results: list of (dataset_name, best_original, best_gaussian, best_clique)

    Mỗi dataset ghi 3 hàng theo thứ tự: Cải tiến → Gốc → CLIQUE.
    Thêm cột sigma cho GCBD Gaussian và cột intervals/threshold cho CLIQUE.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header
    headers = [
        "Dataset", "Algorithm",
        "Số cụm", "Số phần tử được phân cụm",
        "l (no_grid)", "T (max_iters)", "Sigma",
        "Intervals (CLIQUE)", "Threshold (CLIQUE)",
        "Time (s)", "Silhouette", "BSS", "WSS", "WB", "DB", "Dunn", "CH",
        "%Time so với gốc",
    ]

    header_fill = PatternFill("solid", start_color="4472C4")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center")

    thin   = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # Màu xen kẽ dòng
    fill_improved = PatternFill("solid", start_color="DCE6F1")  # xanh nhạt — Gaussian
    fill_original = PatternFill("solid", start_color="FFFFFF")  # trắng      — Gốc
    fill_clique   = PatternFill("solid", start_color="EBF1DE")  # xanh lá nhạt — CLIQUE

    data_font = Font(name="Arial", size=10)

    # Các cột số thực (0-indexed từ col 1): Time=10, Sil=11, BSS=12, WSS=13, WB=14, DB=15, Dunn=16, CH=17, %Time=18
    FLOAT_COLS = {10, 11, 12, 13, 14, 15, 16, 17, 18}

    row_idx = 2
    for entry in all_results:
        dataset_name = entry[0]
        best_orig    = entry[1]
        best_gauss   = entry[2]
        best_clique  = entry[3] if len(entry) > 3 else None

        orig_time = best_orig.get("time") if best_orig else None

        algo_rows = [
            ("Cải tiến", best_gauss,  fill_improved),
            ("Gốc",      best_orig,   fill_original),
            ("CLIQUE",   best_clique, fill_clique),
        ]

        for algo_idx, (algo_label, data, fill) in enumerate(algo_rows):
            is_first = (algo_idx == 0)

            if data is None:
                row_data = [
                    dataset_name if is_first else None,
                    algo_label,
                    None, None, None, None, None, None, None,
                    None, None, None, None, None, None, None, None, None,
                ]
            else:
                time_val = data.get("time")

                # Tham số riêng từng thuật toán
                if algo_label == "CLIQUE":
                    l_display  = None
                    t_val      = None
                    sigma_val  = None
                    intervals  = data.get("intervals")
                    threshold  = data.get("threshold")
                else:
                    l_val      = data.get("no_grid")
                    l_display  = (l_val - 1) if l_val is not None else None
                    t_val      = data.get("max_iters")
                    sigma_val  = data.get("sigma") if algo_label == "Cải tiến" else None
                    intervals  = None
                    threshold  = None

                # %Time so với gốc: chỉ tính cho hàng Cải tiến và CLIQUE
                if algo_label in ("Cải tiến", "CLIQUE") and orig_time:
                    pct_time = time_val / orig_time if time_val is not None else None
                else:
                    pct_time = None

                row_data = [
                    dataset_name if is_first else None,
                    algo_label,
                    data["n_clusters"],
                    data["n_assigned"],
                    l_display,
                    t_val,
                    sigma_val,
                    intervals,
                    threshold,
                    round(time_val, 6) if time_val is not None else None,
                    data["Silhouette"],
                    data["BSS"],
                    data["WSS"],
                    data["WB"],
                    data["DB"],
                    data["Dunn"],
                    data["CH"],
                    round(pct_time, 6) if pct_time is not None else None,
                ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = center
                cell.border    = border
                if col_idx in FLOAT_COLS and val is not None:
                    cell.number_format = "0.000000"

            row_idx += 1

    # Độ rộng cột
    col_widths = [14, 12, 10, 26, 13, 13, 10, 18, 18, 14, 14, 18, 18, 12, 12, 12, 12, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    wb.save(out_path)
    print(f"\n✅ Đã lưu kết quả: {out_path}")


def main():
    os.makedirs(RESULT_DIR, exist_ok=True)

    # Tìm tất cả file .txt trong thư mục dataset
    txt_files = list(Path(DATASET_DIR).glob("*.txt"))
    if not txt_files:
        print(f"⚠️  Không tìm thấy file .txt trong '{DATASET_DIR}'")
        return

    n_combos       = (len(GRID_PARAMS["no_grid"])
                      * len(GRID_PARAMS["percentile"])
                      * len(GRID_PARAMS["max_iters"]))
    n_combos_clique = len(CLIQUE_PARAMS["intervals"]) * len(CLIQUE_PARAMS["threshold"])

    print(f"📦 Tìm thấy {len(txt_files)} dataset(s): {[f.name for f in txt_files]}")
    print(f"GCBD gốc     : {n_combos} tổ hợp / dataset")
    print(f"GCBD Gaussian: {n_combos * len(SIGMA_VALUES)} tổ hợp / dataset ({len(SIGMA_VALUES)} sigma × {n_combos})")
    print(f"CLIQUE       : {n_combos_clique} tổ hợp / dataset (l=[5,50], c=[0,5])")

    all_results = []
    for fpath in sorted(txt_files):
        try:
            dataset_name, X = load_txt_dataset(str(fpath))
        except Exception as e:
            print(f"⚠️  Lỗi khi đọc {fpath.name}: {e}")
            continue

        best = run_best_for_dataset(dataset_name, X)
        all_results.append((
            dataset_name,
            best["original"],
            best["gaussian"],
            best["clique"],
        ))

    if all_results:
        out_path = os.path.join(RESULT_DIR, "results.xlsx")
        write_xlsx(all_results, out_path)
        print(f"📁 Thư mục kết quả: {RESULT_DIR}/")


if __name__ == "__main__":
    main()