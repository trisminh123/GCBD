"""
benchmark_runtime.py
Đo thời gian chạy trung bình của 3 thuật toán: GCBD gốc, GCBD Gaussian, CLIQUE
Chạy N_RUNS lần mỗi dataset, lấy trung bình, rồi vẽ chart so sánh.

Cấu trúc dataset: file .txt tab-separated, cột PoiID/IdPoints, NEAR_X, NEAR_Y

Chạy:
    python benchmark_runtime.py

Kết quả:
    runtime_results/runtime_results.xlsx   ← bảng thời gian
    runtime_results/runtime_chart.png      ← chart so sánh
"""

import os
import time
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from sklearn.preprocessing import MinMaxScaler
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from Algorithm.fastgrid import fastgrid
from Algorithm.fastgrid_gaussian import fastgrid_gaussian

try:
    from pyclustering.cluster.clique import clique as clique_algo
    CLIQUE_AVAILABLE = True
except ImportError:
    CLIQUE_AVAILABLE = False
    print("CẢNH BÁO: pyclustering chưa được cài. Chạy: pip install pyclustering")


# ═══════════════════════════════════════════════════════════════
#  ★  THAM SỐ — TỰ CHỈNH TẠI ĐÂY  ★
# ═══════════════════════════════════════════════════════════════

DATASET_DIR = "Datasets/OSM_dataset"   # Thư mục chứa file .txt
RESULT_DIR  = "results/runtime"      # Thư mục lưu kết quả
N_RUNS      = 50                     # Số lần chạy mỗi thuật toán để lấy trung bình

# --- GCBD gốc ---
GCBD_NO_GRID    = 21      # l: số ô chia lưới mỗi chiều
GCBD_PERCENTILE = 0.1     # percentile ngưỡng mật độ
GCBD_MAX_ITERS  = 5       # T: số vòng lặp tối đa

# --- GCBD Gaussian ---
GAUSS_NO_GRID    = 21     # l
GAUSS_PERCENTILE = 0.1    # percentile
GAUSS_MAX_ITERS  = 5      # T
GAUSS_SIGMA      = 0.5    # σ: độ rộng Gaussian kernel

# --- CLIQUE ---
CLIQUE_INTERVALS  = 20    # l: số khoảng chia mỗi chiều
CLIQUE_THRESHOLD  = 2     # c: ngưỡng mật độ tối thiểu

# --- Chart ---
CHART_TITLE  = "Comparison chart about implementation times"
CHART_XLABEL = "Datasets"
CHART_YLABEL = "Runtime (Seconds)"

# ═══════════════════════════════════════════════════════════════


def load_dataset(filepath: str):
    """Đọc file .txt tab-separated, trả về (tên_dataset, mảng_numpy (n,2))."""
    df = pd.read_csv(filepath, sep="\t")
    df.columns = [c.strip() for c in df.columns]

    id_col = next((c for c in ["PoiID", "IdPoints"] if c in df.columns), None)
    if id_col is None:
        raise ValueError(f"Không tìm thấy cột PoiID/IdPoints trong {filepath}")

    X = df[["NEAR_X", "NEAR_Y"]].values.astype(float)
    return Path(filepath).stem, X


def run_gcbd(X):
    fastgrid(X,
             no_grid    = GCBD_NO_GRID,
             percentile = GCBD_PERCENTILE,
             max_iters  = GCBD_MAX_ITERS)


def run_gaussian(X):
    fastgrid_gaussian(X,
                      no_grid    = GAUSS_NO_GRID,
                      percentile = GAUSS_PERCENTILE,
                      max_iters  = GAUSS_MAX_ITERS,
                      sigma      = GAUSS_SIGMA)


def run_clique(X):
    scaler = MinMaxScaler()
    X_scaled = scaler.fit_transform(X).tolist()
    inst = clique_algo(X_scaled, CLIQUE_INTERVALS, CLIQUE_THRESHOLD)
    inst.process()


def measure_time(fn, X, n_runs: int):
    """Chạy fn(X) n_runs lần, trả về (mean_sec, std_sec, list_times)."""
    times = []
    for _ in range(n_runs):
        t0 = time.perf_counter()
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            fn(X)
        times.append(time.perf_counter() - t0)
    arr = np.array(times)
    return float(arr.mean()), float(arr.std()), times


def benchmark_dataset(dataset_name, X):
    print(f"\n  ▶ {dataset_name}  (n={len(X)})")

    # GCBD gốc
    mean_gcbd, std_gcbd, _ = measure_time(run_gcbd, X, N_RUNS)
    print(f"    GCBD gốc     : {mean_gcbd:.3f}s ± {std_gcbd:.3f}s")

    # GCBD Gaussian
    mean_gauss, std_gauss, _ = measure_time(run_gaussian, X, N_RUNS)
    print(f"    GCBD Gaussian: {mean_gauss:.3f}s ± {std_gauss:.3f}s")

    # CLIQUE
    if CLIQUE_AVAILABLE:
        mean_clique, std_clique, _ = measure_time(run_clique, X, N_RUNS)
        print(f"    CLIQUE       : {mean_clique:.3f}s ± {std_clique:.3f}s")
    else:
        mean_clique = std_clique = None
        print(f"    CLIQUE       : bỏ qua (pyclustering chưa cài)")

    return {
        "dataset"     : dataset_name,
        "gcbd_mean"   : mean_gcbd,
        "gcbd_std"    : std_gcbd,
        "gauss_mean"  : mean_gauss,
        "gauss_std"   : std_gauss,
        "clique_mean" : mean_clique,
        "clique_std"  : std_clique,
    }


def write_xlsx(rows: list, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Runtime"

    headers = [
        "Dataset",
        "GCBD gốc (s)", "GCBD gốc Std",
        "GCBD Gaussian (s)", "GCBD Gaussian Std",
        "CLIQUE (s)", "CLIQUE Std",
    ]

    hfill  = PatternFill("solid", start_color="4472C4")
    hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(border_style="thin", color="999999")
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = center; cell.border = brd

    alt_fills = [
        PatternFill("solid", start_color="DCE6F1"),
        PatternFill("solid", start_color="FFFFFF"),
    ]
    dfont = Font(name="Arial", size=10)

    for ri, row in enumerate(rows, 2):
        fill = alt_fills[ri % 2]
        vals = [
            row["dataset"],
            row["gcbd_mean"],  row["gcbd_std"],
            row["gauss_mean"], row["gauss_std"],
            row["clique_mean"], row["clique_std"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=round(v, 6) if isinstance(v, float) else v)
            cell.font = dfont; cell.fill = fill
            cell.alignment = center; cell.border = brd
            if ci > 1 and v is not None:
                cell.number_format = "0.000000"

    for ci, w in enumerate([16, 16, 14, 18, 16, 12, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    wb.save(out_path)
    print(f"\n✅ Đã lưu Excel: {out_path}")


def draw_chart(rows: list, out_path: str):
    datasets    = [r["dataset"] for r in rows]
    gcbd_times  = [r["gcbd_mean"]  for r in rows]
    gauss_times = [r["gauss_mean"] for r in rows]
    clique_times = [r["clique_mean"] for r in rows]

    x = range(len(datasets))

    fig, ax = plt.subplots(figsize=(12, 6))
    fig.patch.set_facecolor("#F5F5F5")
    ax.set_facecolor("#F5F5F5")

    ax.plot(x, gcbd_times,  color="#2196F3", marker="o", linewidth=2,
            markersize=6, label="GCBD")
    ax.plot(x, gauss_times, color="#F44336", marker="o", linewidth=2,
            markersize=6, label="GK-GCBD")
    if CLIQUE_AVAILABLE and any(v is not None for v in clique_times):
        ax.plot(x, clique_times, color="#4CAF50", marker="s", linewidth=2,
                markersize=6, label="CLIQUE")

    ax.set_title(CHART_TITLE, fontsize=12, fontweight="bold", pad=10)
    ax.set_xlabel(CHART_XLABEL, fontsize=12)
    ax.set_ylabel(CHART_YLABEL, fontsize=12)
    ax.set_xticks(list(x))
    ax.set_xticklabels(datasets, rotation=0, fontsize=12)
    ax.yaxis.set_minor_locator(ticker.AutoMinorLocator())
    ax.grid(True, which="major", linestyle="-",  linewidth=0.5, color="#CCCCCC")
    ax.grid(True, which="minor", linestyle="--", linewidth=0.3, color="#DDDDDD")
    ax.legend(loc="upper left", fontsize=9, framealpha=0.9)

    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"✅ Đã lưu chart : {out_path}")


def main():
    os.makedirs(RESULT_DIR, exist_ok=True)

    txt_files = sorted(Path(DATASET_DIR).glob("*.txt"))
    if not txt_files:
        print(f"⚠️  Không tìm thấy file .txt trong '{DATASET_DIR}'")
        return

    print(f"📦 Tìm thấy {len(txt_files)} dataset(s)")
    print(f"⚙️  Mỗi thuật toán chạy {N_RUNS} lần / dataset → lấy trung bình\n")
    print(f"Tham số GCBD gốc    : no_grid={GCBD_NO_GRID}, percentile={GCBD_PERCENTILE}, max_iters={GCBD_MAX_ITERS}")
    print(f"Tham số GCBD Gauss  : no_grid={GAUSS_NO_GRID}, sigma={GAUSS_SIGMA}, max_iters={GAUSS_MAX_ITERS}")
    print(f"Tham số CLIQUE      : intervals={CLIQUE_INTERVALS}, threshold={CLIQUE_THRESHOLD}")
    print("=" * 60)

    rows = []
    for fpath in txt_files:
        try:
            name, X = load_dataset(str(fpath))
        except Exception as e:
            print(f"⚠️  Lỗi đọc {fpath.name}: {e}")
            continue
        rows.append(benchmark_dataset(name, X))

    if not rows:
        print("⚠️  Không có dataset nào chạy được.")
        return

    xlsx_path  = os.path.join(RESULT_DIR, "runtime_results.xlsx")
    chart_path = os.path.join(RESULT_DIR, "runtime_chart.png")

    write_xlsx(rows, xlsx_path)
    draw_chart(rows, chart_path)

    print(f"\n📁 Xong! Kết quả trong thư mục: {RESULT_DIR}/")


if __name__ == "__main__":
    main()